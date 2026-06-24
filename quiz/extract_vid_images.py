# -*- coding: utf-8 -*-
"""Extract embedded diagrams from VID.xlsx and attach them to questions.json."""
import io
import json
import os
import re

import openpyxl
from PIL import Image

XLSX = "VID.xlsx"
QUESTIONS = "questions.json"
DATA_JS = "data.js"
OUT_DIR = os.path.join("img", "vid")
MIN_DIMENSION = 30

HEADER_CELLS = frozenset(
    {"TEST NUMBER", "QUESTION", "IMAGE", "POSSIBLE ANSWERS", "CORRECT ANSWER"}
)


def is_qtext(val):
    if not val:
        return False
    text = str(val).strip()
    if not text:
        return False
    return text.upper() not in HEADER_CELLS


def normalize_question(text):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", str(text).lower())).strip()


def question_rows(ws, odd_sheet):
    col = 2 if odd_sheet else 1
    rows = []
    for row in range(1, ws.max_row + 1):
        if is_qtext(ws.cell(row, col).value):
            rows.append(row)
    return rows


def anchor_row(ws, image):
    anchor = image.anchor
    kind = type(anchor).__name__
    if kind in ("OneCellAnchor", "TwoCellAnchor"):
        return anchor._from.row + 1
    if kind == "AbsoluteAnchor":
        y = int(anchor.pos.y)
        cumulative = 0.0
        for row in range(1, ws.max_row + 1):
            height = ws.row_dimensions[row].height or 15.0
            cumulative += height * 96.0 / 72.0
            if cumulative > y:
                return row
        return ws.max_row
    return None


def nearest_question_row(image_row, question_row_list):
    candidates = [row for row in question_row_list if row <= image_row]
    return max(candidates) if candidates else None


def image_area(image):
    return image.size[0] * image.size[1]


def image_usable(image):
    return min(image.size) >= MIN_DIMENSION


def save_image(image, path):
    ext = (image.format or "PNG").lower()
    if ext == "jpeg":
        ext = "jpg"
    if ext not in {"png", "jpg", "gif", "webp"}:
        ext = "png"
    path = os.path.splitext(path)[0] + "." + ext
    if ext in {"jpg", "jpeg"} and image.mode in ("RGBA", "P"):
        image = image.convert("RGB")
    image.save(path, quality=92)
    return path


def parse_questions(workbook):
    questions = []
    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        ws = workbook[sheet_name]
        odd_sheet = sheet_index % 2 == 0
        col = 2 if odd_sheet else 1
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if is_qtext(value):
                questions.append(
                    {
                        "sheet": sheet_name,
                        "row": row,
                        "question": re.sub(r"\s+", " ", str(value).strip()),
                    }
                )
    return questions


def build_row_index(workbook):
    row_to_id = {}
    question_id = 0
    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        ws = workbook[sheet_name]
        odd_sheet = sheet_index % 2 == 0
        col = 2 if odd_sheet else 1
        for row in range(1, ws.max_row + 1):
            if is_qtext(ws.cell(row, col).value):
                question_id += 1
                row_to_id[(sheet_name, row)] = question_id
    return row_to_id


def collect_images(workbook, row_to_id):
    best_for_question = {}
    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        ws = workbook[sheet_name]
        odd_sheet = sheet_index % 2 == 0
        qrows = question_rows(ws, odd_sheet)

        for image in getattr(ws, "_images", []):
            image_row = anchor_row(ws, image)
            if image_row is None:
                continue
            question_row = nearest_question_row(image_row, qrows)
            if question_row is None:
                continue
            question_id = row_to_id.get((sheet_name, question_row))
            if not question_id:
                continue

            try:
                pil_image = Image.open(io.BytesIO(image._data()))
            except Exception as exc:
                print("skip image on", sheet_name, "row", image_row, ":", exc)
                continue
            if not image_usable(pil_image):
                continue

            area = image_area(pil_image)
            current = best_for_question.get(question_id)
            if current is None or area > current["area"]:
                best_for_question[question_id] = {
                    "image": pil_image,
                    "area": area,
                    "sheet": sheet_name,
                    "row": question_row,
                }
    return best_for_question


def rewrite_data_js(questions):
    with open(DATA_JS, "w", encoding="utf-8") as handle:
        handle.write("window.QUIZ_DATA = ")
        json.dump(questions, handle, ensure_ascii=False)
        handle.write(";")


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    workbook_data = openpyxl.load_workbook(XLSX, data_only=True)
    workbook = openpyxl.load_workbook(XLSX)

    xlsx_questions = parse_questions(workbook_data)
    quiz_questions = json.load(open(QUESTIONS, encoding="utf-8"))

    if len(xlsx_questions) != len(quiz_questions):
        raise SystemExit(
            "question count mismatch: xlsx=%d json=%d"
            % (len(xlsx_questions), len(quiz_questions))
        )

    mismatches = [
        index + 1
        for index, (left, right) in enumerate(zip(xlsx_questions, quiz_questions))
        if normalize_question(left["question"]) != normalize_question(right["question"])
    ]
    if mismatches:
        raise SystemExit("question text mismatches at ids: %s" % mismatches[:10])

    row_to_id = build_row_index(workbook_data)
    images = collect_images(workbook, row_to_id)

    attached = 0
    replaced = 0
    cleared = 0
    for index, question in enumerate(quiz_questions):
        question_id = index + 1
        payload = images.get(question_id)
        if not payload:
            image_path = question.get("image")
            if image_path and str(image_path).startswith("img/vid/"):
                question["image"] = None
                cleared += 1
            elif image_path and not os.path.exists(image_path):
                question["image"] = None
                cleared += 1
            continue

        rel_path = os.path.join(OUT_DIR, "%03d" % question_id)
        saved = save_image(payload["image"], rel_path)
        rel_path = saved.replace("\\", "/")

        if question.get("image"):
            replaced += 1
        else:
            attached += 1
        question["image"] = rel_path

    json.dump(
        quiz_questions,
        open(QUESTIONS, "w", encoding="utf-8"),
        ensure_ascii=False,
        indent=1,
    )
    rewrite_data_js(quiz_questions)

    with_images = sum(1 for question in quiz_questions if question.get("image"))
    print("xlsx questions:", len(xlsx_questions))
    print("embedded diagrams mapped:", len(images))
    print("newly attached:", attached)
    print("replaced existing image paths:", replaced)
    print("cleared stale non-vid paths:", cleared)
    print("total questions with image:", with_images, "/", len(quiz_questions))
    print("saved under:", OUT_DIR)


if __name__ == "__main__":
    main()
